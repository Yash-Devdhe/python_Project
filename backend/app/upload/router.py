from fastapi import APIRouter, UploadFile, File, Depends, HTTPException
from sqlalchemy.orm import Session
import csv
from io import TextIOWrapper
from openpyxl import load_workbook
from ..deps import get_db, get_current_user
from .. import models
from datetime import datetime

router = APIRouter(prefix="/upload", tags=["upload"])


def _read_csv(file: UploadFile):
    return csv.DictReader(TextIOWrapper(file.file, encoding="utf-8"))


def _read_xlsx(file: UploadFile):
    wb = load_workbook(filename=file.file, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield {headers[i]: row[i] for i in range(len(headers))}


def _iter_rows(file: UploadFile):
    name = (file.filename or "").lower()
    if name.endswith(".csv"):
        yield from _read_csv(file)
    elif name.endswith(".xlsx") or name.endswith(".xls"):
        yield from _read_xlsx(file)
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format. Use .csv or .xlsx")


@router.post("/products")
def upload_products(file: UploadFile = File(...), db: Session = Depends(get_db), user=Depends(get_current_user)):
    created, updated = 0, 0
    for row in _iter_rows(file):
        sku = str(row.get("sku", "")).strip()
        name = (row.get("name") or "").strip()
        category = (row.get("category") or "").strip()
        price = float(row.get("price") or 0)
        if not sku:
            continue
        product = db.query(models.Product).filter(models.Product.sku == sku).first()
        if product:
            product.name = name or product.name
            product.category = category or product.category
            product.price = price or product.price
            updated += 1
        else:
            product = models.Product(sku=sku, name=name, category=category, price=price)
            db.add(product)
            created += 1
    db.commit()
    return {"created": created, "updated": updated}


@router.post("/customers")
def upload_customers(file: UploadFile = File(...), db: Session = Depends(get_db), user=Depends(get_current_user)):
    created, updated = 0, 0
    for row in _iter_rows(file):
        customer_code = str(row.get("customer_id", "")).strip()
        name = (row.get("name") or "").strip()
        email = (row.get("email") or "").strip()
        region = (row.get("region") or "").strip()
        if not customer_code:
            continue
        customer = db.query(models.Customer).filter(models.Customer.customer_id == customer_code).first()
        if customer:
            customer.name = name or customer.name
            customer.email = email or customer.email
            customer.region = region or customer.region
            updated += 1
        else:
            customer = models.Customer(customer_id=customer_code, name=name, email=email, region=region)
            db.add(customer)
            created += 1
    db.commit()
    return {"created": created, "updated": updated}


@router.post("/transactions")
def upload_transactions(file: UploadFile = File(...), db: Session = Depends(get_db), user=Depends(get_current_user)):
    created, skipped = 0, 0
    for row in _iter_rows(file):
        order_id = str(row.get("order_id", "")).strip()
        sku = (row.get("sku") or "").strip()
        customer_code = (row.get("customer_id") or "").strip()
        quantity = int(row.get("quantity") or 0)
        revenue = float(row.get("revenue") or 0)
        order_date_str = (row.get("order_date") or "").strip()
        if not (order_id and sku and customer_code and order_date_str):
            skipped += 1
            continue
        product = db.query(models.Product).filter(models.Product.sku == sku).first()
        customer = db.query(models.Customer).filter(models.Customer.customer_id == customer_code).first()
        if not product or not customer:
            skipped += 1
            continue
        try:
            order_date = datetime.fromisoformat(order_date_str)
        except Exception:
            try:
                order_date = datetime.strptime(order_date_str, "%Y-%m-%d %H:%M:%S")
            except Exception:
                try:
                    order_date = datetime.strptime(order_date_str, "%Y-%m-%d")
                except Exception:
                    skipped += 1
                    continue
        existing = db.query(models.Transaction).filter(models.Transaction.order_id == order_id).first()
        if existing:
            continue
        txn = models.Transaction(
            order_id=order_id,
            product_id=product.id,
            customer_id=customer.id,
            quantity=quantity,
            revenue=revenue,
            order_date=order_date,
        )
        db.add(txn)
        created += 1
    db.commit()
    return {"created": created, "skipped": skipped}
